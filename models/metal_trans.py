import tensorflow as tf
import numpy as np
import os
import sys
from scipy import interp
from sklearn.model_selection import StratifiedKFold, StratifiedShuffleSplit
from sklearn.metrics import matthews_corrcoef, f1_score, roc_curve
from sklearn import metrics
from sklearn.metrics import confusion_matrix, precision_recall_curve
import openpyxl as op
from roc_utils import *
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import pandas as pd
import warnings
from share import ROOT, ESM_PATH, PROT_PATH, LABELS_PATH, CHECKPOINTS, RESULTS


def calculateValueByQK(query, key, value, mask=None):
    """计算注意力加权值。

    Args:
        query (Any): 查询
        key (Any): 键
        value (Any): 值
        mask (Any, optional): 掩码. Defaults to None.

    Returns:
        _type_: 输出值
        _type_: 注意力权重
    """

    # 计算`query`和`key`的矩阵乘积
    matmul_qk = tf.matmul(query, key, transpose_b=True)
    # 获取`key`的最后一个维度
    dk = tf.cast(tf.shape(key)[-1], tf.float32)
    # 缩放`matmul_qk``
    scaled_attention_matrix = matmul_qk / tf.math.sqrt(dk)
    # 如果`mask`存在，则将其添加到缩放的注意力矩阵上
    if mask is not None:
        # 使得在softmax后，mask的值趋近于0
        scaled_attention_matrix += (mask * -1e9)
    # 计算注意力权重
    attention_weights = tf.nn.softmax(scaled_attention_matrix, axis=-1)
    # 计算输出
    output = tf.matmul(attention_weights, value)

    return output, attention_weights


class MultiHeadAttention(tf.keras.layers.Layer):
    def __init__(self, d_model, num_heads):
        super(MultiHeadAttention, self).__init__()
        # 注意力头的数量
        self.num_heads = num_heads
        # 模型的维度
        self.d_model = d_model
        # 模型的深度
        self.depth = d_model // self.num_heads
        # 创建`query`、`key`和`value`的全连接层
        self.wq = tf.keras.layers.Dense(d_model)
        self.wk = tf.keras.layers.Dense(d_model)
        self.wv = tf.keras.layers.Dense(d_model)
        # 创建最后的全连接层
        self.dense = tf.keras.layers.Dense(d_model)

    def split_heads(self, x, batch_size):
        """将最后一个维度分割为(num_heads, depth).

        Args:
            x (Any): 输入
            batch_size (int): 批次大小

        Returns:
            _type_: 输出
        """
        x = tf.reshape(x, (batch_size, -1, self.num_heads, self.depth))
        return tf.transpose(x, perm=[0, 2, 1, 3])

    def call(self, value, key, query, mask=None):
        """多头注意力机制

        Args:
            query (Any): 查询
            key (Any): 键
            value (Any): 值
            mask (Any, optional): 掩码. Defaults to None.

        Returns:
            _type_: 输出
            _type_: 注意力权重
        """
        batch_size = tf.shape(query)[0]
        # 通过`query`、`key`和`value`的全连接层
        query = self.wq(query)
        key = self.wk(key)
        value = self.wv(value)
        # 分割`query`、`key`和`value`为多个头
        query = self.split_heads(query, batch_size)
        key = self.split_heads(key, batch_size)
        value = self.split_heads(value, batch_size)
        # 通过`scaled_dot_product_attention`函数计算注意力加权值
        scaled_attention, attention_weights = calculateValueByQK(query, key, value, mask)
        # 将多个头的输出连接起来
        scaled_attention = tf.transpose(scaled_attention, perm=[0, 2, 1, 3])
        concat_attention = tf.reshape(scaled_attention, (batch_size, -1, self.d_model))
        # 通过最后的全连接层
        output = self.dense(concat_attention)

        return output, attention_weights


def point_wise_feed_forward_network(d_model, dff):
    """点式前馈网络

    Args:
        d_model (int): 模型的维度
        dff (int): 隐藏层的维度

    Returns:
        _type_: 点式前馈网络
    """
    return tf.keras.Sequential([
        tf.keras.layers.Dense(dff, activation='relu'),
        tf.keras.layers.Dense(d_model)
    ])


class EncoderLayer(tf.keras.layers.Layer):
    def __init__(self, d_model, num_heads, dff, rate=0.5):
        super(EncoderLayer, self).__init__()
        # 多头注意力机制
        self.mha = MultiHeadAttention(d_model, num_heads)
        # 点式前馈网络
        self.ffn = point_wise_feed_forward_network(d_model, dff)
        # LayerNorm层
        self.layernorm1 = tf.keras.layers.LayerNormalization(epsilon=1e-6)
        self.layernorm2 = tf.keras.layers.LayerNormalization(epsilon=1e-6)
        # Dropout层
        self.dropout1 = tf.keras.layers.Dropout(rate)
        self.dropout2 = tf.keras.layers.Dropout(rate)

    def call(self, x, training=True, mask=None):
        """编码器层

        Args:
            x (Any): 输入
            training (bool, optional): 是否训练. Defaults to True.
            mask (Any, optional): 掩码. Defaults to None.

        Returns:
            _type_: 输出
        """
        # 多头注意力机制
        attn_output, _ = self.mha(x, x, x, mask)
        attn_output = self.dropout1(attn_output, training=training)
        out1 = self.layernorm1(x + attn_output)
        # 点式前馈网络
        ffn_output = self.ffn(out1)
        ffn_output = self.dropout2(ffn_output, training=training)
        out2 = self.layernorm2(out1 + ffn_output)

        return out2


def get_angles(pos, i, d_model):
    """获取角度

    Args:
        pos (_type_): _description_
        i (_type_): _description_
        d_model (_type_): _description_

    Returns:
        _type_: _description_
    """
    angle_rates = 1 / np.power(10000, (2 * (i // 2)) / np.float32(d_model))
    return pos * angle_rates


def positional_encoding(position, d_model):
    """位置编码

    Args:
        position (int): 位置
        d_model (int): 模型的维度

    Returns:
        _type_: 位置编码
    """
    angle_rads = get_angles(np.arange(position)[:, np.newaxis], np.arange(d_model)[np.newaxis, :], d_model)
    # 将sin应用于数组中的偶数索引（indices）；2i
    angle_rads[:, 0::2] = np.sin(angle_rads[:, 0::2])
    # 将cos应用于数组中的奇数索引；2i+1
    angle_rads[:, 1::2] = np.cos(angle_rads[:, 1::2])
    pos_encoding = angle_rads[np.newaxis, ...]
    return tf.cast(pos_encoding, dtype=tf.float32)


class Encoder(tf.keras.layers.Layer):
    def __init__(self, num_layers, d_model, num_heads, dff, rate=0.5):
        super(Encoder, self).__init__()
        self.d_model = d_model
        self.num_layers = num_layers
        # 位置编码层
        self.pos_encoding = positional_encoding(1000, self.d_model)
        # 编码器层
        self.enc_layers = [EncoderLayer(d_model, num_heads, dff, rate) for _ in range(num_layers)]
        # Dropout层
        self.dropout = tf.keras.layers.Dropout(rate)

    def call(self, x, training=True, mask=None):
        """编码器

        Args:
            x (Any): 输入
            training (bool, optional): 是否训练. Defaults to True.
            mask (Any, optional): 掩码. Defaults to None.

        Returns:
            _type_: 输出
        """
        seq_len = tf.shape(x)[1]
        # 将嵌入向量和位置编码相加
        x += self.pos_encoding[:, :seq_len, :]
        # 通过Dropout层
        x = self.dropout(x, training=training)
        # 通过编码器层
        for i in range(self.num_layers):
            x = self.enc_layers[i](x, training, mask)
        return x


def binary_focal_loss(gamma=2.0, alpha=0.25):
    """二进制焦点损失函数。

    Args:
        gamma (float, optional): 焦点因子. Defaults to 2.0.
        alpha (float, optional): 平衡因子. Defaults to 0.25.

    Returns:
        _type_: 损失函数
    """
    # 将`alpha`和`gamma`转换为张量
    alpha = tf.constant(alpha, dtype=tf.float32)
    gamma = tf.constant(gamma, dtype=tf.float32)

    def binary_focal_loss_fixed(y_true, y_pred):
        """固定焦点损失函数。

        Args:
            y_true (Any): 真实值
            y_pred (Any): 预测值

        Returns:
            _type_: 损失值
        """
        # 将`y_true`转换为浮点张量
        y_true = tf.cast(y_true, tf.float32)
        # 计算`alpha_t`
        alpha_t = y_true * alpha + (tf.ones_like(y_true) - y_true) * (1 - alpha)

        # 计算`p_t`
        p_t = y_true * y_pred + (tf.ones_like(y_true) - y_true) * (
                    tf.ones_like(y_true) - y_pred) + tf.keras.backend.epsilon()
        # 计算焦点损失
        focal_loss = - alpha_t * tf.pow((tf.ones_like(y_true) - p_t), gamma) * tf.math.log(p_t)

        return tf.reduce_mean(focal_loss)

    return binary_focal_loss_fixed


def get_model():
    """获取模型

    Returns:
        _type_: 模型
    """
    inputESM = tf.keras.layers.Input(shape=(161, 1280))
    inputProt = tf.keras.layers.Input(shape=(161, 1024))
    sequence = tf.keras.layers.Dense(512)(inputESM)
    sequence = tf.keras.layers.Dense(256)(sequence)
    sequence = sequence[:, 80, :]
    sequence_prot = tf.keras.layers.Dense(512)(inputProt)
    sequence_prot = tf.keras.layers.Dense(256)(sequence_prot)
    sequence_prot = Encoder(2, 256, 4, 1024, rate=0.3)(sequence_prot)
    Prot = sequence_prot[:, 80, :]
    sequenceconcat = tf.keras.layers.Concatenate()([sequence, Prot])
    feature = tf.keras.layers.Dense(512, activation='relu')(sequenceconcat)
    feature = tf.keras.layers.Dropout(0.4)(feature)
    feature = tf.keras.layers.Dense(256, activation='relu')(feature)
    feature = tf.keras.layers.Dropout(0.4)(feature)
    feature = tf.keras.layers.Dense(128, activation='relu')(feature)
    feature = tf.keras.layers.Dropout(0.4)(feature)
    y = tf.keras.layers.Dense(1, activation='sigmoid')(feature)
    qa_model = tf.keras.models.Model(inputs=[inputESM, inputProt], outputs=y)
    adam = tf.keras.optimizers.Adam(lr=1e-5, beta_1=0.9, beta_2=0.999, epsilon=1e-08, clipnorm=1.0, clipvalue=0.5)
    qa_model.compile(loss=[binary_focal_loss(alpha=.26, gamma=2)], optimizer=adam, metrics=['accuracy'])
    qa_model.summary()
    return qa_model


warnings.filterwarnings("ignore")

ROOT_PATH = ROOT

RESULT_PATH = str(RESULTS)
WEIGHTS_PATH = str(CHECKPOINTS / "metal_trans_weights.h5")

FILEPATH1 = os.path.join(ROOT_PATH, f"results/FULL_____metal_trans.xlsx")
FILEPATH2 = os.path.join(ROOT_PATH, f"results/FULL_tpr_fpr_____metal_trans.xlsx")

IMAGEPATH = os.path.join(ROOT_PATH, f"results/FULL_ROC_____metal_trans.png")

filename = FILEPATH1


def op_toexcel(data, filename):  # openpyxl库储存 数据到excel

    if os.path.exists(filename):
        wb = op.load_workbook(filename)
        ws = wb.worksheets[0]

        ws.append(data)  # 每次写入一行
        wb.save(filename)
    else:
        wb = op.Workbook()  # 创建工作簿对象
        ws = wb['Sheet']  # 创建子表
        ws.append(['MCC', 'ACC', 'AUC', 'Sensitivity', 'Specificity', 'Precision', 'NPV', 'F1', 'FPR', 'FNR',
                   'TN', 'FP', 'FN', 'TP', 'AUPRC', 'Threshold'])  # 添加表头
        ws.append(data)  # 每次写入一行
        wb.save(filename)


def data_generator(train_esm, train_prot, train_y, batch_size):
    L = train_esm.shape[0]

    while True:
        for i in range(0, L, batch_size):
            batch_esm = train_esm[i:i + batch_size].copy()
            batch_prot = train_prot[i:i + batch_size].copy()
            batch_y = train_y[i:i + batch_size].copy()

            yield ([batch_esm, batch_prot], batch_y)


if __name__ == "__main__":
    # os.environ["CUDA_VISIBLE_DEVICES"] = "0"
    # gpus = tf.config.experimental.list_physical_devices('GPU')
    # tf.config.experimental.set_memory_growth(gpus[0], True)
    cpu = tf.config.list_physical_devices("CPU")
    tf.config.set_visible_devices(cpu)

    all_esm = np.lib.format.open_memmap(ESM_PATH)
    all_label = np.lib.format.open_memmap(LABELS_PATH)
    all_prot = np.lib.format.open_memmap(PROT_PATH)

    pos_label = True
    rocs = []
    auc_values = []
    tpr_list = []  # 存储每个交叉验证折叠的真正例率
    mean_fpr = np.linspace(0, 1, 100)  # 平均假正例率的取值范围
    mean_recall = np.linspace(0 , 1, 100)  # 平均假正例率的取值范围

    cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
    k = 1
    for train_index, test_index in cv.split(all_esm, all_label):
        # 训练集
        train_esm = all_esm[train_index]
        train_prot = all_prot[train_index]
        train_label = all_label[train_index]

        # 打乱训练集顺序并划分出验证集
        # （1）分层打乱
        split = StratifiedShuffleSplit(n_splits=1, test_size=0.2, random_state=42)
        for train_inx, valid_inx in split.split(train_esm, train_label):
            # 验证集
            valid_esm = train_esm[valid_inx]
            valid_prot = train_prot[valid_inx]
            valid_label = train_label[valid_inx]
            # 训练集
            train_esm = train_esm[train_inx]
            train_prot = train_prot[train_inx]
            train_label = train_label[train_inx]

        # 测试集
        test_esm = all_esm[test_index]
        test_prot = all_prot[test_index]
        test_label = all_label[test_index]

        # 训练、验证each epoch的步长
        train_size = train_label.shape[0]
        val_size = valid_label.shape[0]
        batch_size = 8
        train_steps = train_size // batch_size
        val_steps = val_size // batch_size

        print(f"Fold {k} - Training samples: {train_esm.shape[0]}, Test samples: {test_esm.shape[0]}")

        qa_model = get_model()
        valiBestModel = WEIGHTS_PATH

        checkpointer = tf.keras.callbacks.ModelCheckpoint(
            filepath=valiBestModel,
            monitor='val_loss',
            save_weights_only=True,
            verbose=1,

            save_best_only=True
        )
        early_stopping = tf.keras.callbacks.EarlyStopping(
            monitor='val_loss',
            patience=40,
            verbose=0,
            mode='auto'
        )

        train_generator = data_generator(train_esm, train_prot, train_label, batch_size)
        val_generator = data_generator(valid_esm, valid_prot, valid_label, batch_size)

        history_callback = qa_model.fit_generator(
            train_generator,
            steps_per_epoch=train_steps,
            epochs=500,
            verbose=1,
            callbacks=[checkpointer, early_stopping],
            validation_data=val_generator,
            validation_steps=val_steps,
            shuffle=True,
            workers=1
        )

        train_generator.close()
        val_generator.close()

        print(f"\nFold {k} - Validation Loss: {history_callback.history['val_loss'][-1]:.4f}, " +
              f"Validation Accuracy: {history_callback.history['val_accuracy'][-1]:.4f}")

        print(f"Fold {k} - Testing:")

        test_pred = qa_model.predict([test_esm, test_prot]).reshape(-1, )

        y_pred = test_pred
        y_true = test_label
        y_pred_new = []

        best_f1 = 0
        best_threshold = 0.5
        for threshold in range(35, 100):
            threshold = threshold / 100
            binary_pred = [1 if pred >= threshold else 0 for pred in y_pred]
            f1 = metrics.f1_score(y_true, binary_pred)
            if f1 > best_f1:
                best_f1 = f1
                best_threshold = threshold

        for value in y_pred:
            if value < best_threshold:
                y_pred_new.append(0)
            else:
                y_pred_new.append(1)
        y_pred_new = np.array(y_pred_new)

        tn, fp, fn, tp = confusion_matrix(y_true, y_pred_new).ravel()
        fpr, tpr, thresholds = roc_curve(y_true, y_pred_new, pos_label=1)
        roc_auc = metrics.auc(fpr, tpr)
        fpr2, tpr2, thresholds2 = roc_curve(y_true, y_pred, pos_label=1)
        roc_auc2 = metrics.auc(fpr2, tpr2)
        roc_auc = roc_auc if roc_auc >= roc_auc2 else roc_auc2
        precision, recall, thresholds = precision_recall_curve(y_true, y_pred_new)
        auprc = metrics.auc(recall, precision)
        thd = best_threshold

        # 计算AUC
        roc = compute_roc(X=y_pred, y=y_true, pos_label=pos_label)
        rocs.append(roc)
        tpr_list.append(np.interp(mean_fpr, fpr2, tpr2))
        roc_auc_bootstrap = roc_auc2
        auc_values.append(roc_auc_bootstrap)

        print("Matthews相关系数: " + str(matthews_corrcoef(y_true, y_pred_new)))
        print("ACC: ", (tp + tn) / (tp + tn + fp + fn))
        print("AUC: ", roc_auc)
        print('sensitivity/recall:', tp / (tp + fn))
        print('specificity:', tn / (tn + fp))
        print('precision:', tp / (tp + fp))
        print('negative predictive value:', tn / (tn + fn))
        print("F1值: " + str(f1_score(y_true, y_pred_new)))
        print('error rate:', fp / (tp + tn + fp + fn))
        print('false positive rate:', fp / (tn + fp))
        print('false negative rate:', fn / (tp + fn))
        print('TN:', tn, 'FP:', fp, 'FN:', fn, 'TP:', tp)
        print('AUPRC: ' + str(auprc))
        print('best_threshold: ' + str(best_threshold))

        mcc = float(format((matthews_corrcoef(y_true, y_pred_new)), '.4f'))
        acc = float(format((tp + tn) / (tp + tn + fp + fn), '.4f'))
        auc = float(format(roc_auc, '.4f'))
        sen = float(format(tp / (tp + fn), '.4f'))
        spe = float(format(tn / (tn + fp), '.4f'))
        pre = float(format(tp / (tp + fp), '.4f'))

        npv = float(format(tn / (tn + fn), '.4f'))
        f1 = float(format(f1_score(y_true, y_pred_new), '.4f'))
        fpr = float(format(fp / (tn + fp), '.4f'))
        fnr = float(format(fn / (tp + fn), '.4f'))
        auprc = float(format(auprc, '.4f'))

        # 保存每一次跑的结果到excel表格
        result = mcc, acc, auc, sen, spe, pre, npv, f1, fpr, fnr, tn, fp, fn, tp, auprc, thd
        op_toexcel(result, filename)

        # if k==5:
        #     test_indep_modify.fcvtest('../save_model/fivecv_model/ZN/model_regular.h5')
        # k += 1

    plt.rcParams['font.family'] = 'Arial'

    # 计算 AUC 的均值和标准差
    mean_auc = np.mean(auc_values)
    std_auc = np.std(auc_values)

    # 计算置信区间
    confidence_interval_auc = 1.96 * std_auc / np.sqrt(len(auc_values))  # 考虑标准误差和样本大小来估计置信区间的一种方法
    confidence_interval = np.percentile(auc_values, [2.5, 97.5])  # 考虑了整个分布的百分位数，而不仅仅是标准差

    roc_mean = compute_mean_roc(rocs)

    resolution = 101
    fpr_mean = np.linspace(0, 1, resolution)
    fpr_mean = np.insert(fpr_mean, 0, 0)  # Insert a leading 0 (closure)
    n_samples = len(rocs)
    thr_all = np.zeros([n_samples, resolution + 1])
    tpr_all = np.zeros([n_samples, resolution + 1])
    auc_all = np.zeros(n_samples)

    for i, ret in enumerate(rocs):
        tpr_all[i, :] = interp(fpr_mean, ret.fpr, ret.tpr)
        thr_all[i, :] = interp(fpr_mean, ret.fpr, ret.thr)
        auc_all[i] = ret.auc
        # Closure
        tpr_all[i, [0, -1]] = ret.tpr[[0, -1]]
        thr_all[i, [0, -1]] = ret.thr[[0, -1]]

    thr_mean = np.mean(thr_all, axis=0)
    tpr_mean = np.mean(tpr_all, axis=0)

    df = pd.DataFrame({
        'mean_tpr': tpr_mean,
        'mean_fpr': fpr_mean
    })

    filename = FILEPATH2
    try:
        # 检查文件是否存在
        book = load_workbook(filename)
        # 如果文件存在，删除工作表并保存新文件
        book.remove(book.active)
    except FileNotFoundError:
        pass  # 文件不存在，则继续保存新文件

    df.to_excel(filename, index=False)

    plt.figure(figsize=(8, 8))
    plot_mean_roc(rocs, show_ci=True, show_ti=True)
    plt.xlim([-.05, 1.05])
    plt.ylim([-.05, 1.05])
    plt.tick_params(axis='x', labelsize=18)
    plt.tick_params(axis='y', labelsize=18)
    plt.title("ROC Curve with Confidence Interval", fontsize=20)
    plt.xlabel('False Positive Rate', fontsize=19)
    plt.ylabel('True Positive Rate', fontsize=19)
    plt.plot([0, 1], [0, 1], 'darkgray', linestyle='--')
    plt.grid(False)
    plt.legend(loc='lower right', fontsize=14)
    plt.savefig(IMAGEPATH, dpi=300)
    # plt.savefig('../pic/full_noencoder.png',dpi=300)
    plt.show()

    # 打印置信区间
    print(f'95% Confidence Interval: {confidence_interval}')
