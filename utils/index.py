from pickle import dump, load
from pathlib import Path
from openpyxl import load_workbook


ROOT = Path(__file__).parent.parent
CACHE = ROOT / 'cache'
if not CACHE.exists():
    CACHE.mkdir()

def dump_cache(data, name):
    """保存数据到缓存文件"""
    with open(CACHE / name, 'wb') as f:
        dump(data, f)

def load_cache(name):
    """从缓存文件中读取数据"""
    with open(CACHE / name, 'rb') as f:
        return load(f)

def load_excel(excel_path: Path):
    """读取Excel文件"""
    wb = load_workbook(excel_path)
    ws = wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    return data

def save_dict_as_fasta(data, path):
    """将字典数据保存为fasta文件"""
    with open(path, 'w') as f:
        for key, value in data.items():
            f.write(f'>{key}\n{value}\n')
